# -*- coding: utf-8 -*-
"""
Created on Fri Jul 31 11:23:18 2020

@author: LY
"""

import os
from moviepy.editor import VideoFileClip
from  openpyxl.workbook  import  Workbook    
import sys
sys.path.append('G:\ECPH_LY\MyPythonProject(python3)')
import Head

class FileCheck():
 
    def __init__(self,file_dir):
        self.file_dir = file_dir
    
    def get_filesize(self,filename):
        u"""
        获取文件大小（M: 兆）
        """
        file_byte = os.path.getsize(filename)
        return self.sizeConvert(file_byte)
 
    def get_file_times(self,filename):
        u"""
        获取视频时长（s:秒）
        """
        clip = VideoFileClip(filename)
        file_time = self.timeConvert(clip.duration)
        return file_time
 
    def sizeConvert(self,size):# 单位换算
        K, M, G = 1024, 1024**2, 1024**3
        if size >= G:
            return str(size/G)+'G Bytes'
        elif size >= M:
            return str(size/M)+'M Bytes'
        elif size >= K:
            return str(size/K)+'K Bytes'
        else:
            return str(size)+'Bytes'
    
    def timeConvert(self,size):# 单位换算
        M, H = 60, 60**2
        if size < M:
            return str(size)+u'秒'
        if size < H:
            return u'%s分钟%s秒'%(int(size/M),int(size%M))
        else:
            hour = int(size/H)
            mine = int(size%H/M)
            second = int(size%H%M)
            tim_srt = u'%s小时%s分钟%s秒'%(hour,mine,second)
            return tim_srt
 
    def get_all_file(self,file_dir):
        u"""
        获取视频下所有的文件
        """
        for root, dirs, files in os.walk(file_dir):  
            return files #当前路径下所有非目录子文件
        
def GetVideolength(data_dir,excel_dir):  
    
    video_dir = data_dir
    Info_file = excel_dir + u'/视频时长统计表.xlsx'
    video_list = Head.WalkPath_MultiType(video_dir,['.mp4'])

    w_write = Workbook()  
    ws_write = w_write.active
    ws_write = w_write.worksheets[0]
    firstline = [u'视频文件',u'时长（分）']
    ws_write.append(firstline) 
    base = 1
    for f in video_list:
        obj = FileCheck(f)
        video_time = obj.get_file_times(f)
        ws_write.cell(base + 1,1,f)
        print (video_time)
        ws_write.cell(base + 1,2,video_time)
        base = base + 1

    w_write.save(Info_file)
    return 'success'

if __name__ == "__main__": 
    
    data_dir = u'G:/ECPH_LY/Data/协助同事/康丽利/超星视频词条加工（丽利）'
    excel_dir = u'G:/ECPH_LY/Data/协助同事/康丽利/超星视频词条加工（丽利）'
    GetVideolength(data_dir,excel_dir)
    

  
