# -*- coding: utf-8 -*-
"""
Created on Mon Sep 21 14:37:46 2020

@author: dbk
"""

import os
import openpyxl
from moviepy.editor import VideoFileClip
from moviepy.editor import *
from moviepy.video.io.preview import preview
from  openpyxl.workbook  import  Workbook    
#import sys
#sys.path.append('G:\ECPH_LY\MyPythonProject(python3)')
#import Head

def WalkPath(src_path,file_type):
    
    required_filelist = []
    for root, dirs, files in os.walk(src_path, topdown=False):
        for name in files:
           
            if os.path.splitext(name)[-1] == file_type:
                required_filelist.append(os.path.join(root, name))
           
            
    return required_filelist

class FileCheck():
 
    def __init__(self,file_dir):
        self.file_dir = file_dir
    
    def get_filesize(self,filename):
        u"""
        获取文件大小（M: 兆）
        """
        file_byte = os.path.getsize(filename)
        return self.sizeConvert(file_byte)
 
    def get_file_times(self,filename):
        u"""
        获取视频时长（s:秒）
        """
        clip = VideoFileClip(filename)
        file_time = self.timeConvert(clip.duration)
        return file_time
 
    def sizeConvert(self,size):# 单位换算
        K, M, G = 1024, 1024**2, 1024**3
        if size >= G:
            return str(size/G)+'G Bytes'
        elif size >= M:
            return str(size/M)+'M Bytes'
        elif size >= K:
            return str(size/K)+'K Bytes'
        else:
            return str(size)+'Bytes'
    
    def timeConvert(self,size):# 单位换算
        M, H = 60, 60**2
        if size < M:
            return str(size)+u'秒'
        if size < H:
            return u'%s分钟%s秒'%(int(size/M),int(size%M))
        else:
            hour = int(size/H)
            mine = int(size%H/M)
            second = int(size%H%M)
            tim_srt = u'%s小时%s分钟%s秒'%(hour,mine,second)
            return tim_srt
 
    def get_all_file(self,file_dir):
        u"""
        获取视频下所有的文件
        """
        for root, dirs, files in os.walk(file_dir):  
            return files #当前路径下所有非目录子文件
        
def GetVideolength(data_dir,excel_dir):  
    
    video_dir = data_dir
    Info_file = excel_dir + u'/视频时长统计表.xlsx'
    video_list = WalkPath(video_dir,'.mp4')

    w_write = Workbook()  
    ws_write = w_write.active
    ws_write = w_write.worksheets[0]
    firstline = [u'视频文件',u'时长（分）']
    ws_write.append(firstline) 
    base = 1
    print('开始写入excel')
    for f in video_list:
       
        obj = FileCheck(f)
        video_time = obj.get_file_times(f)
        ws_write.cell(base + 1,1,f)
        #print (video_time)
        ws_write.cell(base + 1,2,video_time)
        print('正在写入：')
        print(f)
        print (video_time)
        base = base + 1

    w_write.save(Info_file)
    print('写入完成')
    return 'success'

def GetExcelColData(excelfile,col,row):
    #获得excel中某列的数据，从第row行开始取
    wb = openpyxl.load_workbook(excelfile)
    names = wb.sheetnames
    # wb.get_sheet_by_name(name) 已经废弃,使用wb[name] 获取指定工作表
    sheet = wb[names[0]]

    # 获取一列数据, sheet.iter_rows() 获取所有的行
    data_cols = []
    base = 0
    for one_column_data in sheet.iter_rows():
        if base >= row:
            data_cols.append(one_column_data[col].value)
        base = base + 1
        #print(one_column_data[0].value)
    '''
    # 获取一行数据, sheet.iter_cols() 获取所有的列
    data_rows = []
    for one_row_data in sheet.iter_cols():
        data_rows.append(one_row_data[0].value)
        print(one_row_data[0].value, end="\t")
    '''
    return data_cols

def ConvertTime2Second(time_obj):
    return (time_obj.hour+time_obj.minute*60 + time_obj.second)

if __name__ == "__main__": 
    
    data_dir = 'G:/ECPH_LY/Data/协助同事/康丽利/超星视频词条加工（丽利）/1/070.机器学习与人工智能（梁）'
    '''
    moviefile = data_dir + '/' + '机器学习与人工智能（一）.mp4'
    excel_dir = data_dir + '/' + '知识点起止时间.xlsx'
    knowledge_name_list = GetExcelColData(excel_dir,0,1)
    begin_time_list = GetExcelColData(excel_dir,1,1)
    end_time_list = GetExcelColData(excel_dir,2,1)
    #获取剪辑后知识点视频的文件名列表，起止时间，对原视频进行剪辑
    myclip = VideoFileClip(moviefile)

    for i in range(len(knowledge_name_list)):
        myclip_file =data_dir + '/' + knowledge_name_list[i] + '.mp4'
        begin_time = ConvertTime2Second(begin_time_list[i])
        end_time = ConvertTime2Second(end_time_list[i])
        myclip2 = myclip.subclip(begin_time,end_time)
        myclip2.write_videofile(myclip_file,fps = 25)
    '''   
    '''
    #视频拼接
    v1 = VideoFileClip(data_dir + '/1.2.3.“西尔勒中文屋子”实验——“图灵测试”的变体.mp4')
    v2 = VideoFileClip(data_dir + '/1.2.2.“图灵测试”实验——能否从表现来评判智能？.mp4')
    L=[]
    L.append(v2)
    L.append(v1)
    final_clip = concatenate_videoclips(L)
    final_clip.to_videofile(data_dir + '/' + 'all.mp4',fps = 25,remove_temp = False)    
    ''' 
    
    
    