# -*- coding: utf-8 -*-
"""
Created on Fri Apr 21 09:07:21 2017

@author: LiangYan
"""
'''
#为画知识图谱而生成符合echarts规则的数据
'''
import Head
import random
import copy
import numpy as np
import math
from  openpyxl.workbook  import  Workbook    
from  openpyxl.writer.excel  import  ExcelWriter  
from openpyxl.utils import get_column_letter
#from  openpyxl.cell  import  get_column_letter   
from  openpyxl.reader.excel  import  load_workbook
pi = np.pi
           
if __name__ == "__main__":
    '''
    #--------------------------生成json数据，随机指定x,y坐标----------------------------------------
    #读入待连接的excel
    excel_file = u'E:/数据可视化/社科词条库/热词hotlink Excel20170906处理.xlsx'
    itemcn_list = Head.getDataFromExcel(excel_file,0,1,1)
    hotlink_list = Head.getDataFromExcel(excel_file,0,0,1)
    itemcn_list_c = itemcn_list[:]
    hotlink_list_c = hotlink_list[:]
    dst_path = u'E:/数据可视化/社科词条库/社科词条库热词数据.json'
    #Echarts自带颜色盘
    color_list = ['#c23531','#2f4554', '#61a0a8', '#d48265', '#91c7ae','#749f83',\
    '#ca8622', '#bda29a','#6e7074', '#546570', '#c4ccd3']
    
    #边数据生成
    links_info = '"links": ['
    for i in range(len(itemcn_list)):
        links_info = links_info + '{"source": "' + itemcn_list[i] + '","target": "'+ hotlink_list[i] + '"},\n'
    links_info = links_info + ']'
    links_info = links_info.replace('},\n]','}\n]')
    #删除重复结点
    itemcn_list.extend(hotlink_list)
    node_list = list(set(itemcn_list))
    #随机生成结点的x，y坐标
    node_num = len(node_list)
    x_list = [random.randint(-50,50) for x in range(node_num)]
    y_list = [random.randint(-50,50) for x in range(node_num)]

    color_index_list = [random.randint(0,10) for k in range(node_num)]
    d1 = {k:itemcn_list.count(k) for k in set(itemcn_list_c)}
    d2 = {k:hotlink_list.count(k) for k in set(hotlink_list_c)}
    weights = {}
    for key in node_list:
        weight = 0
        if key in d1 :
            weight = d1[key]
        if key in d2:
            weight = weight + d2[key]
        weights[key] = weight
    #随机生成结点的颜色值
    #结点数据生成
    #"x":45,"y":45,"symbolSize":12
    nodes_info = '"nodes":['
    for index in range(len(node_list)):
        node = node_list[index]
        x = x_list[index]
        y = y_list[index]
        w = weights[node]
        color = color_list[color_index_list[index]]
        url = u'image://E:/数据可视化/社科词条库/images/' + node+'.jpg'
        nodes_info = nodes_info + '{"name": "' + node + '","x":'+ str(x) +\
                                    ',"symbol":"' + url +\
                                    '","y":'+ str(y) +',"symbolSize":' + \
                                    str(w) +',"color":"'+ color +'"},\n'
    nodes_info = nodes_info + '],'
    nodes_info = nodes_info.replace('},\n]','}\n]')
    content = nodes_info + '\n'+ links_info
    file_dest = open(dst_path, 'w') 
    file_dest.write('{\n')
    file_dest.write(content.encode('UTF-8'))    
    file_dest.write('\n}')
    file_dest.close()
    '''
    #--------------------------生成社科词条库json数据，力引导模式----------------------------------------
    #读入待连接的excel
    excel_file = u'E:/数据可视化/社科词条库/热词hotlink Excel20170906处理.xlsx'
    url_file = u'E:/数据可视化/社科词条库/所有热词的URL-提交给数据部20170911-下.xlsx'
    url_map = {}
    concept_list = Head.getDataFromExcel(url_file,0,1,1)
    url_list = Head.getDataFromExcel(url_file,0,2,1)
    
    itemcn_list = Head.getDataFromExcel(excel_file,0,0,1)
    hotlink_list = Head.getDataFromExcel(excel_file,0,1,1)
    itemcn_list_c = itemcn_list[:]
    hotlink_list_c = hotlink_list[:]
    dst_path = u'E:/数据可视化/社科词条库/社科词条库热词数据.json'
    #Echarts自带颜色盘
    color_list = ['#c23531','#2f4554', '#61a0a8', '#d48265', '#91c7ae','#749f83',\
    '#ca8622', '#bda29a','#6e7074', '#546570', '#c4ccd3']
    
    #边数据生成
    links_info = '"links": ['
    for i in range(len(itemcn_list)):
        links_info = links_info + '{"source": "' + itemcn_list[i] + '","target": "'+ hotlink_list[i] + '"},\n'
    links_info = links_info + ']'
    links_info = links_info.replace('},\n]','}\n]')   
    #删除重复结点
    itemcn_list.extend(hotlink_list)
    node_list = list(set(itemcn_list))
    #节点url数据生成
    for j in range(len(node_list)):
        concept = node_list[j]
        if concept in concept_list:            
            url = url_list[concept_list.index(concept)]
            url_map[concept] = url
        else:
            print 'can not find the concept'
            break
    url_info = '"URLMap":{'
    for key, value in url_info.items():
        
    #节点权重数据生成
    d1 = {k:itemcn_list_c.count(k) for k in set(itemcn_list_c)}
    d2 = {k:hotlink_list.count(k) for k in set(hotlink_list_c)}
    weights = {}
    for key in node_list:
        weight = 0
        if key in d1 :
            weight = d1[key]
        if key in d2:
            weight = weight + d2[key]
        weights[key] = weight

    #结点数据生成
    nodes_info = '"nodes":['
    for index in range(len(node_list)):
        node = node_list[index]
        w = weights[node]*1.5
        nodes_info = nodes_info + '{"name": "' + node +'","symbolSize":"' + str(w) +'"},\n'
    nodes_info = nodes_info + '],'
    nodes_info = nodes_info.replace('},\n]','}\n]')
    content = nodes_info + '\n'+ links_info
    
    file_dest = open(dst_path, 'w') 
    file_dest.write('{\n')
    file_dest.write(content.encode('UTF-8'))    
    file_dest.write('\n}')
    file_dest.close()
    '''
    #--------------------------读入HOTLINK excel文件数据--------------------------------
    #读入待连接的excel
    excel_file = u'E:/ECPH_LY/Data/社科词条库/test0703/所有对象文件/条目信息.xlsx'
    itemcn_list = Head.getDataFromExcel(excel_file,2,0,1)
    hotlink_list = Head.getDataFromExcel(excel_file,2,1,1)
    #边数据生成
    links = 'links: ['
    for i in range(len(itemcn_list)):
        links = links + '{"source": "' + itemcn_list[i] + '","target": "'+ hotlink_list[i] + '"},\n'
    links = links + ']'      
    #删除重复结点
    itemcn_list.extend(hotlink_list)
    node_list = list(set(itemcn_list))
    #结点数据生成
    data = 'data:['
    for n in node_list:
        data = data + '{name: "' + n + '"},'
    data = data + ']'
    '''   

    '''
    #--------------------读入三版学科群excel文件数据（帮三版办公室张主任做的）------------------------
    #读入待连接的excel
    excel_file = u'E:/数据可视化/学科群建设情况/数学物理化学.xlsx'
    #学科列表
    class_list = Head.getDataFromExcel(excel_file,2,0,1)
    #学科主编列表
    chiefeditor_list = Head.getDataFromExcel(excel_file,2,1,1)
    #学科副主编列表
    subeditor_list = Head.getDataFromExcel(excel_file,2,2,1)
    #预计学科条目数
    class_item_Num_list = Head.getDataFromExcel(excel_file,2,3,1)
    #分支列表
    brunch_list = Head.getDataFromExcel(excel_file,1,1,1)
    #分支主编列表
    bruncheditor_list = Head.getDataFromExcel(excel_file,1,2,1)
    #预计分支条目数
    brunch_item_Num_list = Head.getDataFromExcel(excel_file,1,3,1)
    class_Info_list = []
    brunch_Info_list = []
    for i in range(len(class_list)):
        class_Info = class_list[i] + '\\n' + chiefeditor_list[i] + '\\n' +\
                     subeditor_list[i] + '\\n' + str(int(class_item_Num_list[i]))
        class_Info_list.append(class_Info)
    for j in range(len(brunch_list)):
        if bruncheditor_list[j]:
            brunch_Info = brunch_list[j] + '\\n' + bruncheditor_list[j] + '\\n' +\
                      str(int(brunch_item_Num_list[j]))
        else:
            brunch_Info = brunch_list[j]
        brunch_Info_list.append(brunch_Info)
        
    class_col_list = Head.getDataFromExcel(excel_file,1,0,1)
    brunch_col_list = Head.getDataFromExcel(excel_file,1,1,1)
    class_node_list = []
    brunch_node_list = []
    for row in range(len(class_col_list)):
        if class_col_list[row] in class_list:
            index1 = class_list.index(class_col_list[row])
            class_node_list.append(class_Info_list[index1])
        else:
            class_node_list.append(class_col_list[row])
        if brunch_col_list[row] in class_list:
            print brunch_col_list[row]
            index3 = class_list.index(brunch_col_list[row])
            brunch_node_list.append(class_Info_list[index3])
            print index3
        elif brunch_col_list[row] in brunch_list:
            index2 = brunch_list.index(brunch_col_list[row])
            brunch_node_list.append(brunch_Info_list[index2])
        
        else:
            brunch_node_list.append(brunch_col_list[row])
        
    #边数据生成
    
    links = 'links: ['
    for i in range(len(class_node_list)):
        links = links + '{"source": "' + class_node_list[i] + '","target": "'+ brunch_node_list[i] + '"},\n'
    links = links + ']'      
    #删除重复结点
    class_node_list.extend(brunch_node_list)
    node_list = list(set(class_node_list))
    #结点数据生成
    data = 'data:['
    for n in node_list:
        data = data + '{name: "' + n + '"},'
    data = data + ']'
    '''
